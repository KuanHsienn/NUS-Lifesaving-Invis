"""
Master seed key
---------------
Individual events  — (Event Name, Competitor Name)       one row per athlete
Relay/team events  — (Event Name, '<TEAM> <DIV>')        one row per inst+div
                     e.g. Competitor Name = 'NUS A'

Heat allocation
---------------
- Division order: Y -> B -> A -> O
- Flat list across all divisions, split into equal heats (no per-division imbalance)
- Within each heat: fastest entry -> centre lane (0-indexed), spiral outward
  10-lane pool: [4, 5, 3, 6, 2, 7, 1, 8, 0, 9]
"""

import os
import re
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.helpers import is_team_event

DIV_ORDER          = ["Y", "B", "A", "O"]
DEFAULT_POOL_LANES = 10
MIN_HEAT_SIZE      = 5   # minimum swimmers per heat before merging divisions
MASTER_SEED_COLS   = ["Event Name", "Competitor Name", "Team", "Div", "Best Time", "Year"]

# Set True to show seed times in booklet SEED column; comment out for production
SHOW_SEED_TIMES = True

def parse_timing_to_seconds(value):
    if value is None:
        return None
    if isinstance(value, pd.Timedelta):
        return value.total_seconds() if pd.notna(value) else None
    if isinstance(value, datetime.time):
        return value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return value * 86_400 if 0 < value < 1 else float(value)
    if isinstance(value, str):
        v = value.strip().upper()
        if v in ("", "DNS", "DQ", "DNF", "NS", "NAN"):
            return None
        m = re.match(r"(\d+)\s*min\s+(\d+)\s*s\s+(\d+)\s*ms", v, re.IGNORECASE)
        if m:
            return int(m.group(1)) * 60 + int(m.group(2)) + int(m.group(3)) / 1000
        m = re.match(r"^(\d+):(\d{2})\.(\d+)$", v)
        if m:
            return int(m.group(1)) * 60 + int(m.group(2)) + int(m.group(3)) / (10 ** len(m.group(3)))
        try:
            td = pd.to_timedelta(v)
            if pd.notna(td):
                return td.total_seconds()
        except Exception:
            pass
        try:
            f = float(v)
            return f * 86_400 if 0 < f < 1 else f
        except ValueError:
            pass
    return None


def seconds_to_display(seconds) -> str:
    """Format total seconds as 'M:SS.mmm'. Returns 'NS' when None."""
    if seconds is None or (isinstance(seconds, float) and pd.isna(seconds)):
        return "NS"
    m = int(seconds // 60)
    s = seconds % 60
    return f"{m}:{s:06.3f}"

def load_master_seeds(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        return pd.DataFrame(columns=MASTER_SEED_COLS + ["Best Seconds"])
    try:
        df = pd.read_excel(path, sheet_name="Seeds", dtype=str)
    except Exception as e:
        print(f"  Warning: could not load master seeds — {e}")
        return pd.DataFrame(columns=MASTER_SEED_COLS + ["Best Seconds"])
    df.columns = [str(c).strip() for c in df.columns]
    df["Best Seconds"] = df["Best Time"].apply(parse_timing_to_seconds)
    return df


def save_master_seeds(df: pd.DataFrame, path: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    save_df = df.drop(columns=["Best Seconds"], errors="ignore")[MASTER_SEED_COLS]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        save_df.to_excel(writer, sheet_name="Seeds", index=False)
        ws = writer.sheets["Seeds"]
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(max_len + 2, 12)
    print(f"  Master seed saved -> {path}")

def update_master_seeds_from_results(results_path: str, master_seed_path: str) -> pd.DataFrame:
    """
    Read a completed championship results file and update the master seed.

    Individual events  — key (Event Name, Competitor Name): personal best per athlete.
    Relay/team events  — key (Event Name, '<TEAM> <DIV>'): institution's fastest
                         relay time per event+division, e.g. Competitor Name = 'NUS A'.
                         Multiple rows in the results for the same team+div (e.g. from
                         two relay groups) are all collapsed into one record, keeping
                         whichever time is faster.
    """
    print(f"\n  Reading results: {os.path.basename(results_path)}")
    df_results = pd.read_excel(results_path, sheet_name="Bout Timings")
    df_results.columns = [str(c).strip() for c in df_results.columns]

    year_match = re.search(r"(20\d{2})", os.path.basename(results_path))
    year = year_match.group(1) if year_match else "Unknown"

    def _parse(val):
        try:
            td = pd.to_timedelta(val)
            if pd.notna(td):
                return td.total_seconds()
        except Exception:
            pass
        return parse_timing_to_seconds(val)

    df_results["_seconds"] = df_results["Final Backend Timing"].apply(_parse)
    df_master = load_master_seeds(master_seed_path)

    # existing_lookup: key -> df_master index (for rows already on disk)
    # pending_lookup:  key -> dict reference in new_rows (for rows added this run)
    existing_lookup: dict = {}
    for idx, row in df_master.iterrows():
        k = (str(row.get("Event Name", "")).strip().upper(),
             str(row.get("Competitor Name", "")).strip().upper())
        existing_lookup[k] = idx

    pending_lookup: dict = {}
    new_rows, updated, added = [], 0, 0

    for _, row in df_results.iterrows():
        event_name = str(row.get("Event Name", "")).strip().upper()
        team       = str(row.get("Team", "")).strip().upper()
        div        = str(row.get("Div", "")).strip().upper()
        seconds    = row.get("_seconds")

        if not event_name or div in ("SERC", "NAN", ""):
            continue
        if seconds is None or (isinstance(seconds, float) and pd.isna(seconds)):
            continue

        if is_team_event(event_name):
            # One entry per institution+division — fastest relay time for that team
            comp_name = f"{team} {div}"
        else:
            comp_name = str(row.get("Competitor Name", "")).strip().upper()
            if not comp_name or comp_name == "NAN":
                continue

        key     = (event_name, comp_name)
        display = seconds_to_display(seconds)

        if key in existing_lookup:
            idx = existing_lookup[key]
            try:
                existing = float(df_master.at[idx, "Best Seconds"])
            except (TypeError, ValueError):
                existing = float("inf")
            if seconds < existing:
                df_master.at[idx, "Best Time"]    = display
                df_master.at[idx, "Best Seconds"] = seconds
                df_master.at[idx, "Year"]         = year
                df_master.at[idx, "Team"]         = team
                updated += 1
        elif key in pending_lookup:
            # Same key seen again this run (e.g. two relay groups same inst+div)
            # Update in-place if this result is faster
            nr = pending_lookup[key]
            if seconds < nr["Best Seconds"]:
                nr["Best Time"]    = display
                nr["Best Seconds"] = seconds
                nr["Year"]         = year
                nr["Team"]         = team
        else:
            nr = {
                "Event Name":      event_name,
                "Competitor Name": comp_name,
                "Team":            team,
                "Div":             div,
                "Best Time":       display,
                "Best Seconds":    seconds,
                "Year":            year,
            }
            new_rows.append(nr)
            pending_lookup[key] = nr
            added += 1

    if new_rows:
        df_master = pd.concat([df_master, pd.DataFrame(new_rows)], ignore_index=True)

    div_rank = {d: i for i, d in enumerate(DIV_ORDER)}
    df_master["_div_rank"] = df_master["Div"].map(lambda x: div_rank.get(str(x).upper(), 99))
    df_master = (df_master
                 .sort_values(["Event Name", "_div_rank", "Best Seconds"])
                 .drop(columns=["_div_rank"])
                 .reset_index(drop=True))

    save_master_seeds(df_master, master_seed_path)
    print(f"  Done — {added} new, {updated} updated. Total: {len(df_master)} records.")
    return df_master


# ── Heat allocation ────────────────────────────────────────────────────────────

def get_lane_order(pool_lanes: int = DEFAULT_POOL_LANES) -> list:
    """0-indexed lane order. 10-lane: [4,5,6,3,7,2,8,1,9,0]
    Starts at centre, goes right twice, then alternates left/right."""
    center = (pool_lanes - 1) // 2
    order  = [center]
    r, l   = center + 1, center - 1
    for _ in range(2):
        if r <= pool_lanes - 1 and len(order) < pool_lanes:
            order.append(r); r += 1
    go_left = True
    while len(order) < pool_lanes:
        if go_left:
            if l >= 0: order.append(l); l -= 1
        else:
            if r <= pool_lanes - 1: order.append(r); r += 1
        go_left = not go_left
    return order


def divide_into_heats(swimmers: list, max_lanes: int = DEFAULT_POOL_LANES) -> list:
    """
    Split swimmers (sorted slowest→fastest) into evenly-sized heats of at most
    max_lanes. Extras go to the faster (later) heats so earlier heats are never larger.
    """
    n = len(swimmers)
    if n == 0:
        return []
    num_heats = max(1, (n + max_lanes - 1) // max_lanes)
    base, extras = divmod(n, num_heats)
    heats, idx = [], 0
    for i in range(num_heats):
        size = base + (1 if i >= num_heats - extras else 0)
        heats.append(swimmers[idx: idx + size])
        idx += size
    return heats


def allocate_event_heats(df_seeds: pd.DataFrame,
                         pool_lanes: int = DEFAULT_POOL_LANES) -> tuple:
    """
    Assign heats and lanes for one event.

    Priority order:
    1. Keep Y/B/A divisions in separate heats as much as possible.
    2. Division O fills empty lane slots globally (never its own dedicated heat).
    3. Heats are as equal in size as possible given the above.
    4. Within each division's heats, slower athletes in earlier heats.
       Within each heat, fastest athlete gets centre lane spiralling outward.

    Returns (assignments_list, total_heats).
    """
    def _has_time(s) -> bool:
        v = s.get("Best Seconds")
        return v is not None and not (isinstance(v, float) and pd.isna(v))

    def _sorted_entries(sub) -> list:
        seeded   = sub[sub["Best Seconds"].notna()].sort_values("Best Seconds", ascending=False)
        unseeded = sub[sub["Best Seconds"].isna()]
        return pd.concat([unseeded, seeded], ignore_index=True).to_dict("records")

    def _divide_n(swimmers: list, n: int) -> list:
        """Split swimmers into exactly n lists, extras in the later (faster) heats."""
        if not swimmers:
            return [[] for _ in range(n)]
        base, extras = divmod(len(swimmers), n)
        heats, idx = [], 0
        for i in range(n):
            size = base + (1 if i >= n - extras else 0)
            heats.append(list(swimmers[idx: idx + size]))
            idx += size
        return heats

    # Build sorted lists per division, O kept completely separate
    o_entries:   list = []
    non_o_divs:  list = []   # [(div_char, sorted_entries), ...]
    for div in DIV_ORDER:
        sub = df_seeds[df_seeds["Div"].str.strip().str.upper() == div].copy()
        if sub.empty:
            continue
        entries = _sorted_entries(sub)
        if div == "O":
            o_entries = entries
        else:
            non_o_divs.append((div, entries))

    total = sum(len(e) for _, e in non_o_divs) + len(o_entries)
    if total == 0:
        return [], 0

    optimal_heats = max(1, (total + pool_lanes - 1) // pool_lanes)

    # Global target size per heat (used when filling with O later)
    base_t, extras_t = divmod(total, optimal_heats)
    targets = [base_t + (1 if i >= optimal_heats - extras_t else 0)
               for i in range(optimal_heats)]

    # Determine heat lists from non-O athletes only
    if not non_o_divs:
        # Only O athletes — divide them directly
        heat_lists = _divide_n(o_entries, optimal_heats)
        o_entries  = []   # already placed; skip the fill pass

    else:
        # Minimum heats each non-O division needs independently
        min_h = {div: max(1, (len(entries) + pool_lanes - 1) // pool_lanes)
                 for div, entries in non_o_divs}
        total_min_h = sum(min_h.values())

        if total_min_h <= optimal_heats:
            # Case A: every division gets its own heat block
            # Distribute any extra heats to the division with the most
            # athletes per current heat (minimises the size of each sub-heat).
            n_h = dict(min_h)
            for _ in range(optimal_heats - total_min_h):
                best = max(non_o_divs, key=lambda x: len(x[1]) / n_h[x[0]])[0]
                n_h[best] += 1

            heat_lists = []
            for div, entries in non_o_divs:
                heat_lists.extend(_divide_n(entries, n_h[div]))

        else:
            # Case B: too many divisions for the heat budget hence merge
            # Smart greedy grouping: flush a division group into its own
            # heats only when both the group AND the remaining divisions
            # can fill heats to at least floor(total/optimal_heats) each.
            floor_optimal = total // optimal_heats
            groups:    list = []
            pending:   list = []
            heats_used = 0

            for i, (div, entries) in enumerate(non_o_divs):
                pending.extend(entries)
                is_last   = (i == len(non_o_divs) - 1)
                n         = len(pending)
                n_heats_p = max(1, (n + pool_lanes - 1) // pool_lanes)
                min_size  = n // n_heats_p
                rem_after = sum(len(e) for _, e in non_o_divs[i + 1:])
                heats_after = (max(1, (rem_after + pool_lanes - 1) // pool_lanes)
                               if rem_after else 0)
                rem_min   = rem_after // heats_after if heats_after > 0 else 0

                can_flush = (
                    heats_used + n_heats_p + heats_after <= optimal_heats
                    and (is_last or (
                        min_size >= MIN_HEAT_SIZE
                        and min_size >= floor_optimal
                        and rem_min >= floor_optimal
                    ))
                )
                if can_flush:
                    groups.append(list(pending))
                    pending = []
                    heats_used += n_heats_p

            if pending:
                if groups:
                    groups[-1].extend(pending)
                else:
                    groups.append(pending)

            heat_lists = []
            for group in groups:
                heat_lists.extend(divide_into_heats(group, pool_lanes))

    # Fill every heat with O athletes up to its target size
    o_idx = 0
    for i, heat in enumerate(heat_lists):
        fill = targets[i] - len(heat)
        if fill > 0 and o_idx < len(o_entries):
            take = min(fill, len(o_entries) - o_idx)
            heat.extend(o_entries[o_idx: o_idx + take])
            o_idx += take
    while o_idx < len(o_entries):          # safety: overflow to last heat
        heat_lists[-1].append(o_entries[o_idx])
        o_idx += 1

    # Assign heat numbers and lane positions 
    lane_order = get_lane_order(pool_lanes)
    results    = []

    for global_heat, heat in enumerate(heat_lists, start=1):
        timed   = sorted([(s["Best Seconds"], s) for s in heat if _has_time(s)],
                         key=lambda x: x[0])
        untimed = [s for s in heat if not _has_time(s)]
        ordered = [s for _, s in timed] + untimed

        for lane, sw in sorted(
            [(lane_order[r], sw) for r, sw in enumerate(ordered)],
            key=lambda x: x[0]
        ):
            results.append({
                "Heat No.":        f"H{global_heat}",
                "Lane No.":        lane,
                "Competitor Name": sw.get("Competitor Name", ""),
                "Competitor No.":  sw.get("Competitor No.", ""),
                "Team":            sw.get("Team", ""),
                "Div":             sw.get("Div", ""),
                "Seed Time":       sw.get("Best Time", "NS"),
            })

    return results, len(heat_lists)


# Seed lookup and seed DataFrame builder

def lookup_seed(name: str, df_event: pd.DataFrame) -> tuple:
    """
    Return (best_seconds, display_time) by Competitor Name match.

    Works for both individual athletes (name = athlete name) and relay teams
    (name = '<TEAM> <DIV>', e.g. 'NUS A').
    """
    if df_event.empty or not name.strip():
        return None, "NS"
    match = df_event[
        df_event["Competitor Name"].str.strip().str.upper() == name.strip().upper()
    ]
    if match.empty:
        return None, "NS"
    row  = match.sort_values("Best Seconds").iloc[0]
    secs = row.get("Best Seconds")
    if pd.isna(secs):
        return None, "NS"
    return float(secs), str(row.get("Best Time", seconds_to_display(float(secs))))


def build_seed_df(event_name: str, raw_participants: list,
                   df_master_event: pd.DataFrame) -> pd.DataFrame:
    """
    Build the seed DataFrame consumed by allocate_event_heats().

    Individual events  — one row per athlete; seed time from personal best.
    Relay/team events  — one row per team group (Inst + Div + group code);
                         seed time from the institution's best relay time for
                         that event+division ('<INST> <DIV>' key in master seed).
    """
    _EMPTY = pd.DataFrame(columns=["Competitor Name", "Competitor No.",
                                   "Team", "Div", "Best Seconds", "Best Time"])
    if not raw_participants:
        return _EMPTY

    rows = []

    if is_team_event(event_name):
        teams: dict = {}
        for p in raw_participants:
            key = (p.get("Inst", ""), p.get("Div", ""), p.get("Team", ""))
            if key not in teams:
                teams[key] = {"members": [], "nos": [],
                              "inst": p.get("Inst", ""), "div": p.get("Div", "")}
            teams[key]["members"].append(p.get("Competitor", ""))
            teams[key]["nos"].append(p.get("No.", ""))

        for (inst, div, _), data in teams.items():
            relay_key = f"{inst.strip().upper()} {div.strip().upper()}"
            secs, display = lookup_seed(relay_key, df_master_event)
            rows.append({
                "Competitor Name": " / ".join(data["members"]),
                "Competitor No.":  "\n".join(data["nos"]),
                "Team":            data["inst"],
                "Div":             data["div"],
                "Best Seconds":    secs,
                "Best Time":       display,
            })
    else:
        for p in raw_participants:
            name = p.get("Competitor", "")
            secs, display = lookup_seed(name, df_master_event)
            rows.append({
                "Competitor Name": name,
                "Competitor No.":  p.get("No.", ""),
                "Team":            p.get("Inst", ""),
                "Div":             p.get("Div", ""),
                "Best Seconds":    secs,
                "Best Time":       display,
            })

    return pd.DataFrame(rows) if rows else _EMPTY


# SERC sheet

def write_serc_sheet(ws, raw_participants: list) -> None:
    """Write a simple SERC participant table: S/N, Competitors, No., Inst, Pos."""
    teams: dict = {}
    for p in raw_participants:
        key = (p.get("Inst", ""), p.get("Team", ""))
        if key not in teams:
            teams[key] = {"members": [], "nos": [], "inst": p.get("Inst", "")}
        teams[key]["members"].append(p.get("Competitor", ""))
        teams[key]["nos"].append(p.get("No.", ""))

    if not teams:
        return

    hr = 4
    ws.cell(hr, 1).value  = "S/N"
    ws.merge_cells(f"B{hr}:J{hr}")
    ws.cell(hr, 2).value  = "COMPETITORS"
    ws.cell(hr, 11).value = "NO."
    ws.cell(hr, 12).value = "INST"
    ws.cell(hr, 13).value = "POS"
    for col in [1, 11, 12, 13]:
        ws.cell(hr, col).font      = _font(bold=True, size=12)
        ws.cell(hr, col).alignment = _CENTER
        ws.cell(hr, col).border    = _BORDER
    ws.cell(hr, 2).font      = _font(bold=True, size=12)
    ws.cell(hr, 2).alignment = _CENTER
    _border_cell(ws.cell(hr, 2),  left=True,  right=False, top=True, bottom=True)
    for col in range(3, 10):
        _border_cell(ws.cell(hr, col), left=False, right=False, top=True, bottom=True)
    _border_cell(ws.cell(hr, 10), left=False, right=True,  top=True, bottom=True)
    ws.row_dimensions[hr].height = _ROW_HEIGHT

    for sn, (_, data) in enumerate(teams.items(), start=1):
        dr      = hr + sn
        nos_str = "\n".join(data["nos"])
        ws.cell(dr, 1).value  = sn
        ws.merge_cells(f"B{dr}:J{dr}")
        ws.cell(dr, 2).value  = " / ".join(data["members"])
        ws.cell(dr, 11).value = nos_str
        ws.cell(dr, 12).value = data["inst"]
        ws.cell(dr, 13).value = ""
        ws.cell(dr, 1).font      = _font(bold=True, size=12)
        ws.cell(dr, 1).alignment = _CENTER
        ws.cell(dr, 1).border    = _BORDER
        ws.cell(dr, 2).font      = _font(size=11)
        ws.cell(dr, 2).alignment = _CENTER
        _border_cell(ws.cell(dr, 2),  left=True,  right=False, top=True, bottom=True)
        for col in range(3, 10):
            _border_cell(ws.cell(dr, col), left=False, right=False, top=True, bottom=True)
        _border_cell(ws.cell(dr, 10), left=False, right=True,  top=True, bottom=True)
        for col in [11, 12, 13]:
            ws.cell(dr, col).font      = _font(size=11)
            ws.cell(dr, col).alignment = _CENTER
            ws.cell(dr, col).border    = _BORDER
        num_ids = nos_str.count("\n") + 1
        ws.row_dimensions[dr].height = 20 + 15 * (num_ids - 1)


# Programme booklet generation

_THIN   = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT  = Alignment(horizontal="right",  vertical="center")

_ROW_HEIGHT = 24.75


def _font(bold=False, size=11) -> Font:
    return Font(bold=bold, size=size, name="Calibri")


def _border_cell(cell, left=True, right=True, top=True, bottom=True):
    cell.border = Border(
        left   = _THIN if left   else Side(style=None),
        right  = _THIN if right  else Side(style=None),
        top    = _THIN if top    else Side(style=None),
        bottom = _THIN if bottom else Side(style=None),
    )


def _set_col_widths(ws) -> None:
    ws.column_dimensions["A"].width = 7.44
    for letter in "BCDEFGHIJ":
        ws.column_dimensions[letter].width = 10.11
    ws.column_dimensions["K"].width = 11.33
    ws.column_dimensions["L"].width = 6.33
    ws.column_dimensions["M"].width = 7.55
    ws.column_dimensions["N"].width = 8.00
    ws.column_dimensions["O"].width = 8.00
    ws.column_dimensions["P"].width = 6.00


def _apply_row_border(ws, row: int) -> None:
    ws.cell(row, 1).border = _BORDER
    _border_cell(ws.cell(row, 2),  left=True,  right=False, top=True, bottom=True)
    for col in range(3, 10):
        _border_cell(ws.cell(row, col), left=False, right=False, top=True, bottom=True)
    _border_cell(ws.cell(row, 10), left=False, right=True,  top=True, bottom=True)
    for col in range(11, 17):
        ws.cell(row, col).border = _BORDER


def _write_heat_block(ws, current_row: int, heat_idx: int, total_heats: int,
                      entries: list, pool_lanes: int = DEFAULT_POOL_LANES) -> int:
    """
    Write one heat block. Always prints all pool_lanes rows; unoccupied lanes
    are left blank so the race sheet is ready for manual entry.
    Returns the next available row index.
    """
    entry_by_lane = {e["Lane No."]: e for e in entries}

    # "Finals X of N" label
    c = ws.cell(current_row, 16)
    c.value     = f"Finals {heat_idx} of {total_heats} "
    c.font      = _font(size=12)
    c.alignment = _RIGHT
    ws.row_dimensions[current_row].height = _ROW_HEIGHT
    current_row += 1

    # Header row
    hr = current_row
    ws.cell(hr, 1).value  = "LANE"
    ws.merge_cells(f"B{hr}:J{hr}")
    ws.cell(hr, 2).value  = "COMPETITORS"
    ws.cell(hr, 11).value = "NO."
    ws.cell(hr, 12).value = "INST"
    ws.cell(hr, 13).value = "TEAM"
    ws.cell(hr, 14).value = "SEED"
    ws.cell(hr, 15).value = "TIME"
    ws.cell(hr, 16).value = "POS"
    for col in range(1, 17):
        ws.cell(hr, col).font      = _font(bold=True, size=12)
        ws.cell(hr, col).alignment = _CENTER
    _apply_row_border(ws, hr)
    ws.row_dimensions[hr].height = _ROW_HEIGHT
    current_row += 1

    # Uniform row height across all lanes in this heat
    max_ids = max(
        (str(e["Competitor No."]).count("\n") + 1 for e in entry_by_lane.values()),
        default=1
    )
    lane_row_height = 20 + 15 * (max_ids - 1)

    # Lane rows
    for lane_num in range(pool_lanes):
        dr    = current_row
        entry = entry_by_lane.get(lane_num)

        ws.cell(dr, 1).value  = lane_num
        ws.merge_cells(f"B{dr}:J{dr}")
        ws.cell(dr, 2).value  = entry["Competitor Name"] if entry else ""
        ws.cell(dr, 11).value = entry["Competitor No."]  if entry else ""
        ws.cell(dr, 12).value = entry["Team"]            if entry else ""
        ws.cell(dr, 13).value = entry["Div"]             if entry else ""
        ws.cell(dr, 14).value = (entry["Seed Time"] if SHOW_SEED_TIMES else "") if entry else ""
        ws.cell(dr, 15).value = ""
        ws.cell(dr, 16).value = ""

        ws.cell(dr, 1).font      = _font(bold=True, size=12)
        ws.cell(dr, 1).alignment = _CENTER
        ws.cell(dr, 2).font      = _font(size=11)
        ws.cell(dr, 2).alignment = _CENTER
        for col in range(11, 17):
            ws.cell(dr, col).font      = _font(size=11)
            ws.cell(dr, col).alignment = _CENTER
        _apply_row_border(ws, dr)
        ws.row_dimensions[dr].height = lane_row_height
        current_row += 1

    ws.row_dimensions[current_row].height = _ROW_HEIGHT   # spacer
    return current_row + 1


def _match_event_in_seed(event_name: str, df_master: pd.DataFrame):
    """
    Return the master seed slice for event_name.

    1. Exact match (uppercased).
    2. Recall-based fuzzy match with gender guard (MEN/WOMEN/MIXED must agree).
       Recall = fraction of the target's words found in the candidate; Jaccard
       as tiebreaker. Threshold 0.6.
    """
    if df_master.empty:
        return pd.DataFrame()

    upper = event_name.strip().upper()
    exact = df_master[df_master["Event Name"].str.strip().str.upper() == upper]
    if not exact.empty:
        return exact.copy()

    def _gender(name: str) -> str:
        n = name.strip().upper()
        if n.startswith("WOMEN"): return "WOMEN"
        if n.startswith("MIXED"): return "MIXED"
        if n.startswith("MEN"):   return "MEN"
        return ""

    target_gender = _gender(upper)
    target_words  = set(re.findall(r"\w+", upper))
    best_recall, best_jaccard, best_name = 0.0, 0.0, None

    for name in df_master["Event Name"].dropna().unique():
        if _gender(name) != target_gender:
            continue
        name_words   = set(re.findall(r"\w+", name.upper()))
        intersection = len(target_words & name_words)
        recall  = intersection / len(target_words) if target_words else 0.0
        jaccard = intersection / len(target_words | name_words) if (target_words | name_words) else 0.0
        if recall > best_recall or (recall == best_recall and jaccard > best_jaccard):
            best_recall, best_jaccard, best_name = recall, jaccard, name

    if best_recall >= 0.6 and best_name is not None:
        print(f"    -> Fuzzy matched '{event_name}' to '{best_name}' (recall={best_recall:.2f})")
        return df_master[df_master["Event Name"] == best_name].copy()

    return pd.DataFrame()


def generate_seeded_booklet(master_seed_path: str, participants: dict,
                             output_path: str, event_map: dict = None,
                             pool_lanes: int = DEFAULT_POOL_LANES):
    """
    Generate a seeded programme booklet.

    Parameters
    ----------
    master_seed_path : path to Master_Seed_Timing.xlsx
    participants     : {event_name: [{'Competitor', 'No.', 'Inst', 'Div', 'Team'}, ...]}
    output_path      : destination .xlsx path
    event_map        : {event_code: event_name} from process_registrations()
    pool_lanes       : number of lanes (default 10)
    """
    df_master = load_master_seeds(master_seed_path)
    if not df_master.empty:
        print(f"  Master seed: {len(df_master)} records across "
              f"{df_master['Event Name'].nunique()} events")
    else:
        print("  Master seed is empty — all athletes will be NS")

    participants_ci = {k.strip().upper(): v for k, v in participants.items()}

    def _code_num(code: str) -> int:
        m = re.search(r"\d+", code)
        return int(m.group()) if m else 999
    sorted_events = list(
        sorted(event_map.items(), key=lambda x: _code_num(x[0]))
        if event_map
        else [(f"E{i+1}", name) for i, name in enumerate(participants.keys())]
    )
    wb = Workbook()
    wb.remove(wb.active)

    seeded_count = ns_count = 0

    for event_code, event_name in sorted_events:

        ws = wb.create_sheet(title=event_code.upper())
        title_cell = ws.cell(1, 1)
        title_cell.value = f"{event_code}: {event_name}"
        title_cell.font  = _font(bold=True, size=12)
        for r in range(2, 4):
            ws.row_dimensions[r].height = _ROW_HEIGHT

        raw_participants = participants_ci.get(event_name.strip().upper(), [])
        # SERC is registered under the sheet name "SERC", not the full event name,
        # so the direct lookup above may return [] — fall back to any "serc" key.
        if not raw_participants and ("serc" in event_name.strip().lower() or "simulated" in event_name.strip().lower()):
            raw_participants = next(
                (v for k, v in participants_ci.items() if "serc" in k.lower()), []
            )

        if "serc" in event_name.strip().lower() or "simulated" in event_name.strip().lower():
            write_serc_sheet(ws, raw_participants)
            _set_col_widths(ws)
            continue

        df_master_event   = _match_event_in_seed(event_name, df_master)

        if df_master_event.empty and raw_participants and not df_master.empty:
            print(f"  [NS] No seed match for {event_code}: {event_name!r}")

        seed_df = build_seed_df(event_name, raw_participants, df_master_event)

        if not seed_df.empty:
            ns_in_event   = seed_df[seed_df["Best Seconds"].isna()]
            ns_count     += len(ns_in_event)
            seeded_count += seed_df["Best Seconds"].notna().sum()

            # Diagnostics only for individual events — relay NS is printed separately
            if not is_team_event(event_name) and not ns_in_event.empty and not df_master.empty:
                master_names = set(df_master["Competitor Name"].str.strip().str.upper().dropna())
                for _, row in ns_in_event.iterrows():
                    name   = str(row["Competitor Name"]).strip().upper()
                    reason = ("in master seed but not for this event"
                              if name in master_names
                              else "not in master seed at all")
                    print(f"    NS: {row['Competitor Name']} [{row['Team']} {row['Div']}] — {reason}")

        assignments, total_heats = allocate_event_heats(seed_df, pool_lanes)

        if not assignments:
            _set_col_widths(ws)
            continue

        heats_ordered: list = []
        seen: dict = {}
        for entry in assignments:
            h = entry["Heat No."]
            if h not in seen:
                seen[h] = len(heats_ordered)
                heats_ordered.append((h, []))
            heats_ordered[seen[h]][1].append(entry)

        current_row = 4
        for heat_idx, (_, entries) in enumerate(heats_ordered, start=1):
            current_row = _write_heat_block(ws, current_row, heat_idx, total_heats,
                                            entries, pool_lanes)
        _set_col_widths(ws)

    wb.save(output_path)
    print(f"  Booklet saved -> {output_path}")
    print(f"  Seeded: {seeded_count}  |  NS: {ns_count}")
